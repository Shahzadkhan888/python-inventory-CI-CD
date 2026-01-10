import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_count = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, column=2).value
    inventory = product_list.cell(product_row, column=3).value
    price = product_list.cell(product_row, column=4).value
    product_num = product_list.cell(product_row, column=1).value

    if supplier_name in products_per_supplier:
        current_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_products + 1
    else:
        print("CI/CD: Adding a new supplier to the system...")
        products_per_supplier[supplier_name] = 1
print(products_per_supplier)


    #Count of inventory per supplier

if supplier_name in total_value_per_supplier:
    current_total_value = total_value_per_supplier.get(supplier_name)
    total_value_per_supplier[supplier_name] = current_total_value + inventory * price
else:
    total_value_per_supplier[supplier_name] = inventory * price
print(products_per_supplier)
print(total_value_per_supplier)
#Products less than 10
if inventory < 10:
        products_count[product_num] = inventory

print(products_count)

print("test to check the new build")










